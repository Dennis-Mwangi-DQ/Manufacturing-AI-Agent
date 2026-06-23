"""
Tests for BOM generation.
"""
import os
import tempfile

import pandas as pd
import pytest

from app.models import PartRecord
from app.bom_generator import generate_bom, bom_preview_rows, COLUMNS


def make_test_parts():
    return [
        PartRecord(
            part_id="T1B6-12A501",
            part_name="Fender Panel",
            part_type="SHEET_METAL",
            quantity=1,
            material="High-hardness armour steel",
            material_code="ARMOX-500T",
            thickness_mm=4.0,
            mass_kg=24.3,
            bom_level=1,
            has_bends=True,
            bend_count=2,
            source_filename="B4_Q1-T1B6-12A501-DXF-1.DXF",
            engraving_name="Q1-12A501",
            revision=1,
            assy_code="12500",
            sub_assembly_code="12A500",
        ),
        PartRecord(
            part_id="T1B6-12A502",
            part_name="Bracket",
            part_type="SHEET_METAL",
            quantity=1,
            material=None,
            material_inferred=True,
            low_confidence=True,
            thickness_mm=4.0,
            mass_kg=1.2,
            bom_level=1,
            source_filename="B4_Q1-T1B6-12A502-DXF-1.DXF",
            engraving_name="Q1-12A502",
            revision=1,
            assy_code="12500",
            sub_assembly_code="12A500",
        ),
    ]


def test_bom_generates_files():
    parts = make_test_parts()
    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path, csv_path = generate_bom(parts, "test-session-001", tmpdir)
        assert os.path.exists(xlsx_path)
        assert os.path.exists(csv_path)


def test_bom_row_count():
    parts = make_test_parts()
    with tempfile.TemporaryDirectory() as tmpdir:
        _, csv_path = generate_bom(parts, "test-session-002", tmpdir)
        df = pd.read_csv(csv_path)
        assert len(df) == 2


def test_bom_client_columns():
    parts = make_test_parts()
    with tempfile.TemporaryDirectory() as tmpdir:
        _, csv_path = generate_bom(parts, "test-session-003", tmpdir)
        df = pd.read_csv(csv_path)
        assert list(df.columns) == COLUMNS


def test_bom_client_fields_populated():
    parts = make_test_parts()
    preview = bom_preview_rows(parts)
    row = preview[0]
    assert row["SS ENGRAVING NAME"] == "Q1-12A501"
    assert row["DXF File Name"] == "B4_Q1-T1B6-12A501-DXF-1.DXF"
    assert row["Material"] == "BALLISTIC STEEL"
    assert row["Hardness"] == "500"
    assert row["Thickness (mm)"] == 4.0
    assert row["ASSY"] == "12500"
    assert row["SCOPE OF WORK"] == "L+B"


def test_bom_flags_low_confidence():
    parts = make_test_parts()
    preview = bom_preview_rows(parts)
    assert "LOW_CONFIDENCE" in preview[1]["Flags"]
    assert "INFERRED" in preview[1]["Flags"]


def test_bom_xlsx_has_client_header():
    import openpyxl
    parts = make_test_parts()
    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path, _ = generate_bom(parts, "test-session-006", tmpdir)
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["BOM"]
        assert ws.cell(row=1, column=1).value == "T1B6_12500_DXF_BOM"
        assert ws.cell(row=2, column=1).value == "SR NO."
        assert ws.cell(row=3, column=2).value == "12A500"
