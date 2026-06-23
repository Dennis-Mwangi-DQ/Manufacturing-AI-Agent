"""
BOM generator — creates Excel (.xlsx) and CSV outputs from enriched PartRecord list.

Column layout follows the client T1B6 DXF BOM convention (SR NO., SS ENGRAVING NAME,
DXF File Name, Material, Hardness, Thickness, ASSY, SCOPE OF WORK, etc.).
"""
from __future__ import annotations

import csv
import json
import logging
import os
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import PartRecord

logger = logging.getLogger(__name__)

# Colour constants
HEADER_BG = "1F4E79"
HEADER_FG = "FFFFFF"
LOW_CONF_BG = "FFFF00"
FILENAME_BG = "FCE4D6"  # light orange — client highlights engraving / DXF name
GROUP_BG = "D9E1F2"

COLUMNS = [
    "SR NO.",
    "IMAGE",
    "SS ENGRAVING NAME",
    "Revision",
    "DXF File Name",
    "Material",
    "Hardness",
    "Thickness (mm)",
    "Quantity",
    "ASSY",
    "SCOPE OF WORK",
    "Notes",
    "Flags",
]

SCOPE_LEGEND = [
    ("L", "Laser Cutting"),
    ("B", "Bending"),
    ("M", "Machining"),
    ("T", "Tapping"),
    ("CS", "Counter Sunk"),
]

_FILENAME_LETTER_CLIENT_MATERIAL = {
    "B": ("BALLISTIC STEEL", 450),
    "M": ("MILD STEEL", None),
}


def _load_materials() -> list[dict]:
    mat_path = Path(__file__).resolve().parent.parent / "data" / "materials.json"
    if mat_path.exists():
        with open(mat_path, encoding="utf-8") as f:
            return json.load(f).get("materials", [])
    return []


def _lookup_material(code: Optional[str], materials: list[dict]) -> Optional[dict]:
    if not code:
        return None
    for mat in materials:
        if mat.get("code") == code:
            return mat
    return None


def _flags(part: PartRecord) -> str:
    flags = []
    if part.low_confidence:
        flags.append("LOW_CONFIDENCE")
    if part.material_inferred:
        flags.append("INFERRED")
    if part.has_bends:
        flags.append(f"BENDS:{part.bend_count}")
    return "; ".join(flags)


def _scope_of_work(part: PartRecord) -> str:
    """Derive client scope codes from geometry (L/B only — M/T/CS need drawing review)."""
    codes: list[str] = []
    if part.part_type == "SHEET_METAL" or part.source_filename:
        codes.append("L")
    if part.has_bends:
        codes.append("B")
    return "+".join(codes)


def _client_material(part: PartRecord, materials: list[dict]) -> str:
    mat = _lookup_material(part.material_code, materials)
    if mat and mat.get("client_material_name"):
        name = str(mat["client_material_name"])
    elif part.material:
        name = part.material
    else:
        fmeta_letter = None
        if part.source_filename:
            from app.cad_parser import parse_filename_metadata
            fmeta_letter = parse_filename_metadata(part.source_filename).get("material_letter")
        if fmeta_letter and fmeta_letter in _FILENAME_LETTER_CLIENT_MATERIAL:
            name = _FILENAME_LETTER_CLIENT_MATERIAL[fmeta_letter][0]
        else:
            return ""

    if part.material_inferred and name:
        return f"{name} (INFERRED)"
    return name


def _hardness(part: PartRecord, materials: list[dict]) -> str:
    mat = _lookup_material(part.material_code, materials)
    if mat and mat.get("hardness") is not None:
        return str(mat["hardness"])

    if part.source_filename:
        from app.cad_parser import parse_filename_metadata
        letter = parse_filename_metadata(part.source_filename).get("material_letter")
        if letter and letter in _FILENAME_LETTER_CLIENT_MATERIAL:
            h = _FILENAME_LETTER_CLIENT_MATERIAL[letter][1]
            if h is not None:
                return str(h)
    return ""


def _dxf_file_name(part: PartRecord) -> str:
    if part.source_filename:
        return part.source_filename
    if part.source_path:
        return Path(part.source_path).name
    return ""


def _group_key(part: PartRecord) -> str:
    return part.sub_assembly_code or part.parent_assembly or ""


def _sorted_parts(parts: list[PartRecord]) -> list[PartRecord]:
    return sorted(parts, key=lambda p: (_group_key(p), p.part_id))


def _build_part_row(idx: int, part: PartRecord, materials: list[dict]) -> list:
    thickness = part.thickness_mm
    return [
        idx,
        "",
        part.engraving_name or part.part_id,
        part.revision if part.revision is not None else "",
        _dxf_file_name(part),
        _client_material(part, materials),
        _hardness(part, materials),
        thickness if thickness is not None else "",
        part.quantity,
        part.assy_code or "",
        _scope_of_work(part),
        part.notes or "",
        _flags(part),
    ]


def _build_grouped_rows(parts: list[PartRecord], materials: list[dict]) -> list[tuple[str, list]]:
    """Return [(group_code|'', row_values), ...] with group header markers."""
    sorted_p = _sorted_parts(parts)
    output: list[tuple[str, list]] = []
    current_group: Optional[str] = None
    sr = 0

    for part in sorted_p:
        group = _group_key(part)
        if group and group != current_group:
            output.append(("__GROUP__", [group]))
            current_group = group
            sr = 0
        sr += 1
        output.append(("__DATA__", _build_part_row(sr, part, materials)))

    return output


def _flat_data_rows(parts: list[PartRecord], materials: list[dict]) -> list[list]:
    """CSV-friendly flat rows (no group header rows)."""
    rows: list[list] = []
    current_group: Optional[str] = None
    sr = 0
    for part in _sorted_parts(parts):
        group = _group_key(part)
        if group != current_group:
            current_group = group
            sr = 0
        sr += 1
        rows.append(_build_part_row(sr, part, materials))
    return rows


def bom_preview_rows(parts: list[PartRecord]) -> list[dict]:
    """Return BOM rows as dicts for API/UI preview (same columns as Excel/CSV)."""
    materials = _load_materials()
    return [dict(zip(COLUMNS, row)) for row in _flat_data_rows(parts, materials)]


def _bom_title(parts: list[PartRecord]) -> str:
    for part in parts:
        if part.assy_code:
            return f"T1B6_{part.assy_code}_DXF_BOM"
    return "CAD_AGENT_DXF_BOM"


def generate_bom(parts: list[PartRecord], session_id: str, output_dir: str) -> tuple[str, str]:
    """
    Build BOM from enriched PartRecord list.
    Returns (xlsx_path, csv_path).
    """
    os.makedirs(output_dir, exist_ok=True)
    xlsx_path = os.path.join(output_dir, "BOM.xlsx")
    csv_path = os.path.join(output_dir, "BOM.csv")
    materials = _load_materials()

    flat_rows = _flat_data_rows(parts, materials)
    grouped = _build_grouped_rows(parts, materials)

    # ---- Excel (client-style layout) ----
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    title = _bom_title(parts)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal="center")

    header_row = 2
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    header_font = Font(bold=True, color=HEADER_FG, size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ws.row_dimensions[header_row].height = 30

    low_conf_fill = PatternFill(start_color=LOW_CONF_BG, end_color=LOW_CONF_BG, fill_type="solid")
    filename_fill = PatternFill(start_color=FILENAME_BG, end_color=FILENAME_BG, fill_type="solid")
    group_fill = PatternFill(start_color=GROUP_BG, end_color=GROUP_BG, fill_type="solid")
    normal_font = Font(size=10)

    excel_row = header_row + 1
    for kind, row_data in grouped:
        if kind == "__GROUP__":
            cell = ws.cell(row=excel_row, column=2, value=row_data[0])
            cell.font = Font(bold=True, size=10)
            cell.fill = group_fill
            excel_row += 1
            continue

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font = normal_font
            cell.alignment = Alignment(vertical="center")
            if col_idx in (3, 5):
                cell.fill = filename_fill
            flags = row_data[-1] if row_data else ""
            if isinstance(flags, str) and "LOW_CONFIDENCE" in flags:
                cell.fill = low_conf_fill
        excel_row += 1

    # Scope-of-work legend (client key)
    legend_col = len(COLUMNS) + 2
    ws.cell(row=header_row, column=legend_col, value="Key").font = Font(bold=True, size=10)
    for i, (code, label) in enumerate(SCOPE_LEGEND, start=1):
        ws.cell(row=header_row + i, column=legend_col, value=code).font = Font(bold=True, size=10)
        ws.cell(row=header_row + i, column=legend_col + 1, value=label).font = Font(size=10)

    # Totals
    total_row = excel_row
    ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True, size=10)
    ws.cell(row=total_row, column=9, value=sum(p.quantity for p in parts)).font = Font(bold=True, size=10)
    ws.cell(row=total_row, column=3, value=f"{len(parts)} parts total").font = Font(italic=True, size=10)

    col_widths = [len(c) + 2 for c in COLUMNS]
    for row_data in flat_rows:
        for i, val in enumerate(row_data):
            col_widths[i] = min(50, max(col_widths[i], len(str(val)) + 2))
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A3"
    wb.save(xlsx_path)
    logger.info("BOM Excel saved to %s", xlsx_path)

    # ---- CSV ----
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(COLUMNS)
        writer.writerows(flat_rows)
    logger.info("BOM CSV saved to %s", csv_path)

    return xlsx_path, csv_path
